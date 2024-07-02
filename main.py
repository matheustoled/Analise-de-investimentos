#Requests e web scraping
import requests
from bs4 import BeautifulSoup
import json
#Manipulação de planilhas
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl import load_workbook
from openpyxl.styles import Alignment
#Criação de interface
from tkinter import *
import tkinter as tk
#Matemática
import math

#Definição de variáveis iniciais
max_linhas = 1

#Carregar planilha
wb = openpyxl.load_workbook('Investimentos.xlsx', data_only=True)
sheet1 = wb['Sheet1']
sheet1 = wb.active

#Lista com ações que o usuário escolheu
lista_acoes = []

#Definindo cores
preto = "#000000" #black
branco = "#f1ebeb" #white
azul = "#24c0eb" #blue
cinza = "#cacaca" #gray

#Define cores para a formatação condicional
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

#Cria a janela principal
janela = tk.Tk()
janela.title("Análise de Ações")
janela.geometry("375x185")

#Cria um frame para a entrada e o botão
frame = tk.Frame(janela)
frame.pack(pady=20, padx=20, fill='both', expand=True)

#Cria um campo de entrada dentro do frame
entrada = tk.Entry(frame, width=30)
entrada.pack(pady=10)

#Criando função de entrada da informação
def adicionar_acao():
    global max_linhas
    acao = entrada.get()
    if acao:  # Verifica se o campo de entrada não está vazio
        lista_acoes.append(acao)
        max_linhas += 1
        entrada.delete(0, tk.END)  # Limpa o campo de entrada após adicionar o nome
        print(f"Ação '{acao}' adicionado(a)!")
    else:
        print("O campo de entrada está vazio. Por favor, digite uma ação.")

#Cria um botão dentro do frame para enviar o nome
botao = tk.Button(frame, text="Adicionar", command=adicionar_acao, bg=azul, fg=branco, font=("Uvy 13 bold"), relief=RAISED, overrelief=RIDGE)
botao.pack(pady=10)

#Funçao para pegar as informações de cada ativo
def obter_dados_ativos(ativo):

    #Formatar a URL corretamente com o ativo que o usuário escolheu
    url = "https://statusinvest.com.br/acoes/{}".format(ativo)

    #Definir os cabeçalhos HTTP para imitar um navegador real
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    #Fazer request do site para puxar as infos
    requisicao = requests.get(url, headers=headers)

    #ler todo o conteúdo HTML do Status Invest
    site = BeautifulSoup(requisicao.content, "html.parser")

    infos = site.find_all('strong', {'class': 'value d-block lh-4 fs-4 fw-700'})
    infos2 = site.find_all('strong', {'class': 'value'})
    try:
        preco_atual = float(site.find('strong', {'class': 'value'}).text.replace(',','.').strip(''))
    except ValueError:
        preco_atual = 0.0
    try:
        margem_liquida = float(infos[23].text.replace(',','.').replace('%','').strip(''))
    except ValueError:
        margem_liquida = 0.0
    try:
        div_liquida_patrimonio = float(infos[14].text.replace(',','.').strip(''))
    except ValueError:
        div_liquida_patrimonio = 0.0
    try:
        roic = float(infos[26].text.replace(',','.').replace('%','').strip(''))
    except ValueError:
        roic = 0.0
    try:
        div_liquida_ebitda = float(infos[15].text.replace(',','.').strip(''))
    except ValueError:
        div_liquida_ebitda = 0.0
    try:
        tag_along = float(infos2[6].text.replace(',','.').replace('%','').strip(''))
    except ValueError:
        tag_along = 0.0
    try:
        pl = float(infos[1].text.replace(',','.').strip(''))
    except ValueError:
        pl = 0.0
    try:
        p_vp = float(infos[3].text.replace(',','.').strip(''))
    except ValueError:
        p_vp = 0.0
    try:
        dy = float(infos[0].text.replace(',','.').replace('%','').strip(''))
    except ValueError:
        dy = 0.0
    try:
        liq_corrente = float(infos[19].text.replace(',','.').strip(''))
    except ValueError:
        liq_corrente = 0.0
    try:
        roe = float(infos[24].text.replace(',','.').replace('%','').strip(''))
    except ValueError:
        roe = 0.0
    try:
        ev_ebitda = float(infos[4].text.replace(',','.').strip(''))
    except ValueError:
        ev_ebitda = 0.0
    try:
        lpa = float(infos[10].text.replace(',','.').strip(''))
    except ValueError:
        lpa = 0.0
    try:
        vpa = float(infos[8].text.replace(',','.').strip(''))
    except ValueError:
        vpa = 0.0
        
    #Definindo valores padrão
    teto9 = valor_justo = ey = ey2 = 0

    if dy != 0 and preco_atual != 0:
        teto9 = (dy * preco_atual) / 9
    
    if lpa != 0 and vpa != 0:
        try:
            valor_justo = math.sqrt(22.5 * lpa * vpa)
        except ValueError:
            valor_justo = 0

    if lpa != 0 and preco_atual != 0:
        ey = (lpa / preco_atual) * 100
    
    if vpa != 0 and preco_atual != 0:
        ey2 = (vpa / preco_atual) * 100
    
    # Lista com os indicadores padrões de investimento
    lista_indicadores = {
        'Ativo': ativo,
        'Valor': preco_atual,
        'VPA': vpa,
        'Teto 9%': teto9,
        'Valor Justo por Ação': valor_justo,
        'Margem Líquida': margem_liquida,
        'LPA': lpa,
        'DY': dy,
        'EY': ey,
        'EY2': ey2,
        'P/L': pl,
        'P/VP': p_vp,
        'ROE': roe,
        'ROIC': roic,
        'Tag Along': tag_along,
        'EV/EBITDA': ev_ebitda,
        'Dívida Líquida/Patrimônio': div_liquida_patrimonio,
        'Dívida Líquida/Ebitida': div_liquida_ebitda,
        'Liq. Corrente': liq_corrente
    }
    return lista_indicadores

#Função para formatar as colunas com R$
#def formatar_coluna_como_reais(planilha, coluna):
#    for i in range(2, max_linhas + 1):
#        celula = planilha[f'{coluna}{i}']
#        celula.number_format = 'R$ #,#0.#0'

#Função para formatar as colunas com %
#def formatar_coluna_como_porcentagem(planilha, coluna):
#    for i in range(2, max_linhas + 1):
#        celula = planilha[f'{coluna}{i}']
#        celula.number_format = '0.00%'

#Funções formatação condicional
def formatacao_condicional_bom(planilha, coluna, parametro, valor):
    area = "{}2:{}{}".format(coluna,coluna,max_linhas)
    formatacao_bom = CellIsRule(operator=parametro, formula=[valor], fill=PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'))
    planilha.conditional_formatting.add(area, formatacao_bom)

def formatacao_condicional_bom_b(planilha, coluna, parametro, valori, valorf):
    area = "{}2:{}{}".format(coluna,coluna,max_linhas)
    formatacao_bom = CellIsRule(operator=parametro, formula=[valori,valorf], fill=PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'))
    planilha.conditional_formatting.add(area, formatacao_bom)

def formatacao_condicional_ruim(planilha, coluna, parametro, valor):
    area = "{}2:{}{}".format(coluna,coluna,max_linhas)
    formatacao_ruim = CellIsRule(operator=parametro, formula=[valor], fill=PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'))
    planilha.conditional_formatting.add(area, formatacao_ruim)

def formatacao_condicional_ruim_b(planilha, coluna, parametro, valori, valorf):
    area = "{}2:{}{}".format(coluna,coluna,max_linhas)
    formatacao_ruim = CellIsRule(operator=parametro, formula=[valori,valorf], fill=PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'))
    planilha.conditional_formatting.add(area, formatacao_ruim)

def formatacao_condicional_nulo(planilha, coluna, parametro, valor):
    area = "{}2:{}{}".format(coluna,coluna,max_linhas)
    formatacao_nulo = CellIsRule(operator=parametro, formula=[valor], fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'))
    planilha.conditional_formatting.add(area, formatacao_nulo)


#Função para salvar os dados em Excel
def salvar_dados_excel():
    
    #Lista para armazenar os dados de cada ativo separadamente
    dados_ativos = []

    #Obter os dados para cada ativo
    for i in lista_acoes:
        i = i.strip()  #Remover espaços em branco extras
        try:
            dados = obter_dados_ativos(i)
            dados_ativos.append(dados)
        except Exception as e:
            print(f"Erro ao obter dados para o ativo {i}: {e}")

    #Criar um DataFrame a partir do dicionário
    df = pd.DataFrame(dados_ativos)

    #Exportar o DataFrame para um arquivo Excel
    df.to_excel('Investimentos.xlsx', index=False)
    print("Dados salvos!")

    # Recarregar a planilha para aplicar a formatação
    wb = openpyxl.load_workbook('Investimentos.xlsx')
    sheet1 = wb.active

    #Formatando números
    #formatar_coluna_como_reais(planilha=sheet1, coluna='B')
    #formatar_coluna_como_porcentagem(planilha=sheet1, coluna='H')
    #formatar_coluna_como_porcentagem(planilha=sheet1, coluna='F')
    #formatar_coluna_como_porcentagem(planilha=sheet1, coluna='M')
    #formatar_coluna_como_porcentagem(planilha=sheet1, coluna='N')
    #formatar_coluna_como_porcentagem(planilha=sheet1, coluna='O')

    #Formatação condicional
    formatacao_condicional_bom(sheet1, 'F', 'greaterThanOrEqual', '10')
    formatacao_condicional_ruim_b(sheet1, 'F', 'between', '0.001', '9.999')
    formatacao_condicional_nulo(sheet1, 'F', 'equal', '')

    formatacao_condicional_bom_b(sheet1, 'Q', 'between', '0.001','0.999')
    formatacao_condicional_ruim(sheet1, 'Q', 'greaterThanOrEqual', '1')
    formatacao_condicional_nulo(sheet1, 'Q', 'equal', '')

    formatacao_condicional_bom(sheet1, 'N', 'greaterThanOrEqual', '15')
    formatacao_condicional_ruim_b(sheet1, 'N', 'between', '0.001', '14.999')
    formatacao_condicional_nulo(sheet1, 'N', 'equal', '')

    formatacao_condicional_bom_b(sheet1, 'R', 'between', '0.001','2.999')
    formatacao_condicional_ruim(sheet1, 'R', 'greaterThanOrEqual', '3')
    formatacao_condicional_nulo(sheet1, 'R', 'equal', '')

    formatacao_condicional_bom(sheet1, 'O', 'equal', '100')
    #formatacao_condicional_ruim_b(sheet1, 'O', 'between', '0.001', '99.999')
    formatacao_condicional_nulo(sheet1, 'O', 'equal', '')

    formatacao_condicional_bom_b(sheet1, 'K', 'between', '0.001','9.999')
    formatacao_condicional_ruim(sheet1, 'K', 'greaterThanOrEqual', '10')
    formatacao_condicional_nulo(sheet1, 'K', 'equal', '')

    formatacao_condicional_bom_b(sheet1, 'L', 'between', '0.001','1.499')
    formatacao_condicional_ruim(sheet1, 'L', 'greaterThanOrEqual', '1.5')
    formatacao_condicional_nulo(sheet1, 'L', 'equal', '')

    formatacao_condicional_bom(sheet1, 'S', 'greaterThanOrEqual', '1')
    formatacao_condicional_ruim_b(sheet1, 'S', 'between', '0.001', '0.999')
    formatacao_condicional_nulo(sheet1, 'S', 'equal', '')

    formatacao_condicional_bom(sheet1, 'M', 'greaterThanOrEqual', '16')
    formatacao_condicional_ruim_b(sheet1, 'M', 'between', '0.001', '15.999')
    formatacao_condicional_nulo(sheet1, 'M', 'equal', '')

    formatacao_condicional_bom_b(sheet1, 'P', 'between', '0.001','4.999')
    formatacao_condicional_ruim(sheet1, 'P', 'greaterThanOrEqual', '5')
    formatacao_condicional_nulo(sheet1, 'P', 'equal', '')

    formatacao_condicional_bom(sheet1, 'I', 'greaterThanOrEqual', '20')
    formatacao_condicional_ruim_b(sheet1, 'I', 'between', '0.001', '19.999')
    formatacao_condicional_nulo(sheet1, 'I', 'equal', '')

    #Congelar coluna A
    sheet1.freeze_panes = "B1"

    #Salvar planilha
    wb.save('Investimentos.xlsx')

#Cria um botão para salvar os dados em Excel
botao_salvar = tk.Button(frame, text="Salvar Dados", command=salvar_dados_excel, bg=azul, fg=branco, font=("Uvy 13 bold"), relief=RAISED, overrelief=RIDGE)
botao_salvar.pack(pady=10)

#Executa o loop principal da interface
janela.mainloop()